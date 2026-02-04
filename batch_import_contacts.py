"""
Batch import 92K contacts from all_contacts.xlsx into RDS contacts table.
Uses streaming Excel read + batch inserts of 500 rows to avoid memory/lag issues.
"""

import os
import uuid
import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

DB_CONFIG = {
    "host": os.getenv("DB_HOST"),
    "port": os.getenv("DB_PORT", 5432),
    "dbname": os.getenv("DB_NAME"),
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD"),
    "sslmode": "require",
}

EXCEL_PATH = "./public/all_contacts.xlsx"
BATCH_SIZE = 500

INSERT_SQL = """
INSERT INTO contacts
    (first_name, last_name, full_name, phone, email, dob,
     address_line_1, city, state_county, postal_code,
     source, sales_signature_token)
VALUES %s
"""


def parse_phone(val):
    """Convert numeric phone to string."""
    if val is None:
        return None
    return str(int(val)) if isinstance(val, (int, float)) else str(val).strip()


def clean(val):
    """Strip whitespace, convert 'nan'/empty to None."""
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in ("nan", "none", ""):
        return None
    return s


def parse_dob(val):
    """Convert DD/MM/YYYY to YYYY-MM-DD for PostgreSQL."""
    s = clean(val)
    if s is None:
        return None
    try:
        parts = s.split("/")
        if len(parts) == 3:
            return f"{parts[2]}-{parts[1]}-{parts[0]}"  # YYYY-MM-DD
    except Exception:
        pass
    return s  # return as-is if already in correct format


def build_row(row_values):
    """Map Excel row to DB tuple."""
    # Excel cols: [index, first_name, last_name, phone, email, dob, county, city, postcode, street_address]
    _, first_name, last_name, phone, email, dob, county, city, postcode, street_address = row_values

    first_name = clean(first_name)
    last_name = clean(last_name)
    full_name = " ".join(filter(None, [first_name, last_name])) or None
    phone = parse_phone(phone)
    email = clean(email)
    dob_str = parse_dob(dob)
    county = clean(county)
    city = clean(city)
    postcode = clean(postcode)
    street_address = clean(street_address)

    return (
        first_name,
        last_name,
        full_name,
        phone,
        email,
        dob_str,
        street_address,     # -> address_line_1
        city,
        county,             # -> state_county
        postcode,           # -> postal_code
        "Bulk Import",
        str(uuid.uuid4()),  # sales_signature_token
    )


def main():
    print("Connecting to database...", flush=True)
    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = False
    cur = conn.cursor()

    print(f"Reading {EXCEL_PATH} in streaming mode...", flush=True)
    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active

    batch = []
    total_inserted = 0
    errors = 0

    rows = ws.iter_rows(min_row=2, values_only=True)  # skip header

    for row_values in rows:
        try:
            record = build_row(row_values)
            batch.append(record)
        except Exception as e:
            errors += 1
            if errors <= 5:
                print(f"  Skipping bad row: {e}", flush=True)
            continue

        if len(batch) >= BATCH_SIZE:
            try:
                execute_values(cur, INSERT_SQL, batch, page_size=BATCH_SIZE)
                conn.commit()
                total_inserted += len(batch)
                print(f"  Inserted {total_inserted} rows...", flush=True)
            except Exception as e:
                conn.rollback()
                print(f"  ERROR on batch at row ~{total_inserted}: {e}", flush=True)
                errors += len(batch)
            batch = []

    # Insert remaining rows
    if batch:
        try:
            execute_values(cur, INSERT_SQL, batch, page_size=BATCH_SIZE)
            conn.commit()
            total_inserted += len(batch)
        except Exception as e:
            conn.rollback()
            print(f"  ERROR on final batch: {e}", flush=True)
            errors += len(batch)

    wb.close()
    cur.close()
    conn.close()

    print(f"\nDone! Inserted: {total_inserted} | Errors: {errors}", flush=True)


if __name__ == "__main__":
    main()
